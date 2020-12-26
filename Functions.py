from datetime import datetime
from multiprocessing import connection

Cno = '0'  # 用于最后的注销


# 注册账号密码
# 对应参数分别是学号,密码,密保
def register(studentNumber, password, SecretGuard):
    import pymysql
    db = pymysql.connect("localhost", "root", "123456", "onecartoon")
    # 使用cursor()方法获取操作游标
    # 使用cursor()方法获取操作游标
    cursor = db.cursor()
    ch = "select 学号 from 学生基本信息 where 学号='%s'" % studentNumber
    cursor.execute(ch)
    ch1 = cursor.fetchone()
    if ch1 == None:
        print("学号输入错误")
    else:
        sql = """INSERT INTO 注册(学号,
                      密码,密保)
                      VALUES ('%s', '%s', '%s')""" % (studentNumber, password, SecretGuard)
        try:
            # 执行sql语句
            cursor.execute(sql)
            # 提交到数据库执行
            db.commit()
        except:
            # 如果发生错误则回滚
            db.rollback()
            # 关闭数据库连接
            db.close()
        else:
            print("注册成功！")
        return


# 登录
# 需要对应的变量来存放登录的账号
def load(studentNumber, password):
    global Cno
    import pymysql.cursors
    connection = pymysql.connect(host='localhost', port=3306, user='root', password='123456', db='onecartoon',
                                 charset='utf8mb4', cursorclass=pymysql.cursors.DictCursor)
    cur = connection.cursor()
    # pattern = re.compile()
    sqlDL = "select 学号  from 注册 where 注册.学号='%s' and 注册.密码='%s'" % (studentNumber, password)
    cur.execute(sqlDL)
    jilu = cur.fetchone()  # 注册表中没有对应账号和密码
    if jilu == None:
        return False
    elif jilu != None:
        Cno = studentNumber
        return True


# 信息的录入
# 对应参数分别是 学号 性别 姓名 学院 专业 手机号 余额
def Imformation(studentNumber, name, gender, College, major, mobileNumber):
    import pymysql
    # 打开数据库连接
    db = pymysql.connect("localhost", "root", "123456", "onecartoon")
    # 使用cursor()方法获取操作游标
    cursor = db.cursor()
    money = 100
    lable = [studentNumber, name, gender, College, major, mobileNumber]
    lable1 = ['学号', '姓名', '性别', '学院', '专业', '手机号']
    str = ''  # 初始化为空串
    str1 = ''
    flag = 0
    for i in range(4, 6):
        # print(i)
        if lable[i] != '':
            str1 += lable1[i] + ","
            flag = 1
            str += "'" + lable[i] + "',"
    str = str[:-1]
    str1 = str1[:-1]
    # print(studentNumber)
    # # SQL 插入语句
    if flag == 0:
        sql = "INSERT INTO 学生基本信息(学号,姓名,性别,学院,余额) VALUES('%s','%s','%s','%s',%s)" % (
        studentNumber, name, gender, College, money)
    else:
        sql = "INSERT INTO 学生基本信息(学号,姓名,性别,学院,%s,余额) VALUES ('%s','%s','%s','%s',%s, %s)" % (
            str1, studentNumber, name, gender, College, str, money)
    try:
        # 执行sql语句
        print(sql)
        cursor.execute(sql)
        # 提交到数据库执行
        db.commit()
    except:
        # 如果发生错误则回滚
        print("信息录入错误")
        db.rollback()
        # 关闭数据库连接
        db.close()
    else:
        print('信息录入成功')
        return


# 对应参数是 属性和具体的值   比如性别，男
def Check(fun, natural):
    import pymysql.cursors
    connection = pymysql.connect(host='localhost', port=3306, user='root', password='123456', db='onecartoon',
                                 charset='utf8mb4', cursorclass=pymysql.cursors.DictCursor)
    cur = connection.cursor()
    if fun == '性别':
        sqlJiaoYi = "select * from 交易记录 where 性别='%s'" % (natural)
    elif fun == '学号':
        sqlJiaoYi = "select * from 交易记录 where 学号='%s'" % (natural)
    elif fun == '消费类型':
        sqlJiaoYi = "select * from 交易记录 where 消费类型='%s'" % (natural)
    elif fun == '1':
        sqlJiaoYi = "select * from 交易记录"
    cur.execute(sqlJiaoYi)
    result = cur.fetchall()
    for data in result:
        print(data)
    cur.close()
    connection.close()


# Check('性别', '女')
# 充值校园卡   要改
# 对应参数是 钱数 学号 时间
def Recharge(money, nameNumber):
    import pymysql
    db = pymysql.connect("localhost", "root", "123456", "onecartoon")
    # 使用cursor()方法获取操作游标
    cursor = db.cursor()
    time = datetime.now()
    ch = "select 学号 from 学生基本信息 where 学号='%s'" % (nameNumber)
    cursor.execute(ch)  # 加载到数据库中
    ch1 = cursor.fetchone()  # 将数据拿出来
    if ch1 is None:
        return 0
    else:
        sql1 = "select 学号,余额 from 学生基本信息 where 学号='%s'" % (nameNumber)
        cursor.execute(sql1)
        ch3 = cursor.fetchone()
        print(ch3)
        sql = "update 学生基本信息 set 余额=余额+%s where 学号='%s'" % (money, nameNumber)
        sql3 = "INSERT INTO 交易记录(学号,消费类型,金额变动,消费时间,余额) VALUES ('%s','充值',%s,'%s',%s)" % (ch3[0], money, time, ch3[1] + float(money))
        print(sql)
        print(sql3)
        try:
            # 执行sql语句
            cursor.execute(sql)
            # 提交到数据库执行
            db.commit()
            cursor.execute(sql3)
            db.commit()
            return 2
        except:
            # 如果发生错误则回滚
            db.rollback()
            # 关闭数据库连接
            db.close()
            return 1


#  修改信息
# 对应参数是学号
def Exchange(name):
    import pymysql.cursors
    connection = pymysql.connect(host='localhost', port=3306, user='root', password='123456', db='onecartoon',
                                 charset='utf8mb4', cursorclass=pymysql.cursors.DictCursor)
    cur = connection.cursor()
    secretGuard = input("请输入密保：")
    sqlXH = "select 密保 from 注册 where 注册.学号='%s'" % (name)
    cur.execute(sqlXH)
    jilu = cur.fetchone()
    # for data in jilu:
    if jilu == {'密保': secretGuard}:
        NewpPassword = input("请输入新密码：")
        sqlNewPassword = "update 注册 set 密码='%s' where 学号='%s'" % (NewpPassword, name)
        print("修改成功")
    elif jilu != {'密码': secretGuard}:
        print("密保错误")
        Exchange(name)
    cur.close()
    connection.close()


# 基本信息表的输出
# 对应参数是要输出的表 比如要输出学生基本信息  str=学生基本信息
def inputInformation(str):
    import pymysql.cursors
    connection = pymysql.connect(host='localhost', port=3306, user='root', password='123456', db='onecartoon',
                                 charset='utf8mb4', cursorclass=pymysql.cursors.DictCursor)
    try:
        cursor = connection.cursor()
        sql = "select * from %s" % (str)
        cursor.execute(sql)
        result = cursor.fetchall()
        # 信息的输出
        for data in result:
            print(data)
    except Exception as e:
        print("信息输入错误")
        # 提示错误信息
        # print(e)
        cursor.close()
        connection.close()
    return


# str = input("请输入所要输出的表的名字：")
# inputInformation(str)


# 注销账号
def logOut():
    import pymysql
    db = pymysql.connect("localhost", "root", "123456", "onecartoon")
    # 使用cursor()方法获取操作游标
    # 使用cursor()方法获取操作游标
    cursor = db.cursor()
    sql = "delete from 学生基本信息 where 学号='%s'" % Cno
    sql1 = "delete from 注册 where 学号='%s'" % Cno
    sql2 = "delete from 交易记录 where 学号='%s'" % Cno
    try:
        # 执行sql语句
        cursor.execute(sql1)
        # 提交到数据库执行
        db.commit()
        cursor.execute(sql2)
        db.commit()
        cursor.execute(sql)
        db.commit()
    except:
        # 如果发生错误则回滚
        print("注销失败！")
        db.rollback()
        # 关闭数据库连接
        db.close()
    else:
        print("注销成功！")


# 学生基本信息查询
def informationInput(studentNumber, name, gender, College, major, mobileNumber):
    lable = [studentNumber, name, gender, College, major, mobileNumber]
    lable1 = ['学号', '姓名', '性别', '学院', '专业', '手机号']
    str = ""  # 初始化为空串
    flag = 0

    for i in range(6):
        if lable[i] != '':
            flag = 1
            # str += (lable1[i] + "='" + lable[i] + "' and ")
            str += (lable1[i] + " LIKE '%" + lable[i] + "%' and ")

    str = str[:-5]
    # print(str)

    import pymysql.cursors
    connection = pymysql.connect(host='localhost', port=3306, user='root', password='123456', db='onecartoon',
                                 charset='utf8mb4', cursorclass=pymysql.cursors.DictCursor)
    try:
        cursor = connection.cursor()
        if flag == 0:
            sql = "select * from 学生基本信息"
        else:
            sql = "select * from 学生基本信息 where %s" % str
        cursor.execute(sql)
        result = cursor.fetchall()
        # 信息的输出
        return result
    except Exception as e:
        print("信息输入错误")
        # 提示错误信息
        print(e)
        cursor.close()
        connection.close()


def DataUpdate(studentNumber, attribute, val):
    import pymysql
    db = pymysql.connect("localhost", "root", "123456", "onecartoon")
    # 使用cursor()方法获取操作游标
    # 使用cursor()方法获取操作游标
    cursor = db.cursor()
    sql = "update 学生基本信息 set %s='%s' where 学号='%s'" % (attribute, val, studentNumber)
    print(sql)
    try:
        # 执行sql语句
        cursor.execute(sql)
        # 提交到数据库执行
        db.commit()
    except:
        # 如果发生错误则回滚
        print("修改失败！")
        db.rollback()
        # 关闭数据库连接
        db.close()
    else:
        print("修改成功！")


def DeleteData(studentNumber):
    import pymysql
    db = pymysql.connect("localhost", "root", "123456", "onecartoon")
    # 使用cursor()方法获取操作游标
    # 使用cursor()方法获取操作游标
    cursor = db.cursor()
    # print(studentNumber)
    sql = "DELETE FROM 学生基本信息 WHERE 学号='%s'" % (studentNumber)
    sql1 = "DELETE FROM 交易记录 WHERE 学号='%s'" % (studentNumber)
    try:
        # 执行sql语句
        cursor.execute(sql1)
        # 提交到数据库执行
        db.commit()
        cursor.execute(sql)
        # 提交到数据库执行
        db.commit()
    except:
        # 如果发生错误则回滚
        print("删除信息失败！")
        #         db.rollback()
        #         # 关闭数据库连接
        db.close()
    else:
        print("删除成功！")


# 交易记录输出
def TransactionRecords(studentNumber, name, ConsumptionType):
    lable = [studentNumber, name, ConsumptionType]
    lable1 = ['学号', '姓名', '消费类型']
    str = ""  # 初始化为空串
    flag = 0
    for i in range(3):
        if lable[i] != "":
            flag = 1
            if i == 0:
                str += "学生基本信息." + lable1[i] + " LIKE '%" + lable[i] + "%' and "
            elif i == 1:
                str += "学生基本信息." + lable1[i] + " LIKE '%" + lable[i] + "%' and "
            elif i == 2:
                str += lable1[i] + " LIKE '%" + lable[i] + "%' and "

    if str.__len__()>5:
        str = str[:-5]
    import pymysql.cursors
    connection = pymysql.connect(host='localhost', port=3306, user='root', password='123456', db='onecartoon')
    cur = connection.cursor()
    if flag == 0:
        sql = "select 学生基本信息.学号,学生基本信息.姓名,消费类型,消费时间,金额变动,交易记录.余额 from 学生基本信息,交易记录 where 学生基本信息.学号=交易记录.学号"
    else:
        sql = "select 学生基本信息.学号,学生基本信息.姓名,消费类型,消费时间,金额变动,交易记录.余额 from 交易记录,学生基本信息 where 学生基本信息.学号=交易记录.学号 and %s" % str
    # print(sql)
    cur.execute(sql)
    result = cur.fetchall()
    # print(result)
    # for data in result:
    #     print(data)
    # 更改字段
    list = []
    for i in range(result.__len__()):
        temp = result[i]
        list.append([])
        for j in range(0, 6):
            list[i].append(temp[j])
    return list
    cur.close()
    connection.close()

# 检查密保问题
def SecurityQuestion(name, secretGuard):
    import pymysql.cursors
    connection = pymysql.connect(host='localhost', port=3306, user='root', password='123456', db='onecartoon',
                                 charset='utf8mb4', cursorclass=pymysql.cursors.DictCursor)
    cur = connection.cursor()
    sqlXH = "select 密保 from 注册 where 注册.学号='%s'" % name
    cur.execute(sqlXH)
    jilu = cur.fetchone()
    # for data in jilu:
    cur.close()
    connection.close()
    if jilu == {'密保': secretGuard}:
        return True
    else:
        return False


# 重置密码
def ResetPassword(name, new_password):
    import pymysql
    db = pymysql.connect("localhost", "root", "123456", "onecartoon")
    # 使用cursor()方法获取操作游标
    cursor = db.cursor()
    sqlNewPassword = "update 注册 set 密码='%s' where 学号='%s'" % (new_password, name)
    cursor.execute(sqlNewPassword)  # 提交到数据库执行
    db.commit()
    # 关闭数据库连接
    db.close()


def StudentInformation(path, start, last):
    import pymysql
    from openpyxl import load_workbook
    from datetime import date, datetime
    db = pymysql.connect("localhost", "root", "123456", "onecartoon")
    # 使用cursor()方法获取操作游标
    # 使用cursor()方法获取操作游标
    cursor = db.cursor()
    print("666")
    if path == "":
        path = "student.xlsx"
    workbook = load_workbook(filename=".\\" + path)
    # print(workbook.sheetnames)
    sheet1 = workbook["sheet1"]
    print("999")
    if start == "":
        start = 1
    if last == "":
        last = 1
    start = int(start)
    last = int(last)
    print("start:" + str(start) + ",  end:" + str(last))
    list1 = []
    # from builtins import range
    for q in range(start, last+1):
        list1.append([])
        # print(q)
        for row in sheet1.iter_rows(min_row=q, max_row=q):
            # print(row)
            for cell in row:
                # from builtins import str
                string = str(cell.value)
                # print(string)
                # print(string)
                list1[q - start].append(string)
    print(list1)
    for i in range(0, len(list1)):
        date = list1[i]
        # print(i, date)
        # print(date)
        str1 = ""

        for j in range(6):
            str1 += "'" + date[j] + "',"

        str1 = str1[:-1]
        sql_str = "INSERT INTO 学生基本信息(学号, 姓名, 性别, 学院, 专业, 手机号, 余额) VALUES(%s,%s)" % (str1, int(date[6]))
        print(sql_str)
        try:
            # 执行sql语句
            db.ping(reconnect=True)
            cursor.execute(sql_str)
            db.commit()
        except Exception as e:
            # 如果发生错误则回滚
            print("修改失败！")
            print(e)
            db.rollback()
            # 关闭数据库连接
            db.close()
